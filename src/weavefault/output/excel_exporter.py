"""
WeaveFault ExcelExporter - export FMEA documents to .xlsx format.
"""

from __future__ import annotations

import logging
from pathlib import Path

from weavefault.ingestion.schema import FMEADocument
from weavefault.output.fmea_formatter import FMEAFormatter
from weavefault.standards import (
    build_high_risk_label,
    is_high_risk_row,
    load_standard_profile,
)

logger = logging.getLogger(__name__)

COLOR_HIGH = "FFFF0000"
COLOR_MEDIUM = "FFFFA500"
COLOR_LOW = "FF90EE90"
COLOR_HEADER = "FF2C3E50"
COLOR_WHITE = "FFFFFFFF"


class ExcelExporter:
    """Export FMEADocument to a styled .xlsx file."""

    def __init__(self, standard: str = "IEC_60812") -> None:
        self.standard_profile = load_standard_profile(standard)
        self.formatter = FMEAFormatter(standard=standard)

    def export(self, doc: FMEADocument, output_path: str) -> Path:
        """Export the FMEA document to an Excel workbook."""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        dest = self._resolve_path(output_path, doc.id, ".xlsx")
        dest.parent.mkdir(parents=True, exist_ok=True)

        workbook = openpyxl.Workbook()
        self._write_fmea_sheet(workbook, doc, PatternFill, Font, Alignment)
        self._write_summary_sheet(workbook, doc, Font)

        workbook.save(str(dest))
        logger.info("Excel exported to %s", dest)
        return dest

    def _write_fmea_sheet(self, workbook, doc, PatternFill, Font, Alignment) -> None:
        ws = workbook.active
        ws.title = "FMEA"

        rows = self.formatter.format_document(doc)
        if not rows:
            ws.append(["No FMEA rows generated."])
            return

        headers = list(self.formatter.get_header_labels().values())
        columns = list(rows[0].keys())

        ws.append(headers)
        header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
        header_font = Font(color=COLOR_WHITE, bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        rpn_col_idx = columns.index("rpn") + 1 if "rpn" in columns else None
        for source_row, row_data in zip(doc.rows, rows):
            ws.append([row_data.get(col, "") for col in columns])
            data_row = ws.max_row
            if rpn_col_idx:
                rpn_cell = ws.cell(row=data_row, column=rpn_col_idx)
                rpn = int(rpn_cell.value or 0)
                if is_high_risk_row(
                    source_row,
                    self.standard_profile,
                    threshold=doc.high_risk_threshold,
                ):
                    rpn_cell.fill = PatternFill("solid", fgColor=COLOR_HIGH)
                    rpn_cell.font = Font(bold=True, color=COLOR_WHITE)
                elif rpn >= self.standard_profile.medium_risk_threshold:
                    rpn_cell.fill = PatternFill("solid", fgColor=COLOR_MEDIUM)
                else:
                    rpn_cell.fill = PatternFill("solid", fgColor=COLOR_LOW)

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    def _write_summary_sheet(self, workbook, doc, Font) -> None:
        ws = workbook.create_sheet(title="Summary")
        high_risk_label = build_high_risk_label(
            self.standard_profile,
            doc.high_risk_threshold,
        )
        ws.append(["WeaveFault FMEA Summary"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Document ID", doc.id])
        ws.append(["Version", doc.version])
        ws.append(["Domain", doc.domain])
        ws.append(["Standard", doc.standard])
        ws.append(["Generated At", str(doc.generated_at)])
        ws.append(["Model Used", doc.model_used])
        ws.append([])
        ws.append(["Total Components", doc.total_components])
        ws.append(["Total Failure Modes", len(doc.rows)])
        ws.append([f"High Risk ({high_risk_label})", doc.high_risk_count])
        if doc.rows:
            top = sorted(doc.rows, key=lambda row: row.rpn, reverse=True)[:3]
            ws.append([])
            ws.append(["Top 3 Highest RPN"])
            for row in top:
                ws.append([row.component_name, row.failure_mode, row.rpn])

    @staticmethod
    def _resolve_path(output_path: str, doc_id: str, suffix: str) -> Path:
        dest = Path(output_path)
        if dest.is_dir():
            return dest / f"fmea_{doc_id}{suffix}"
        if not dest.suffix:
            return dest.with_suffix(suffix)
        return dest
